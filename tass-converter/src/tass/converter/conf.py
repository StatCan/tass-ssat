import openpyxl
from pathlib import Path


def convert(path):
    """
    Helper tool to convert xlsx config file to a in-memory json equivalent.
    """

    conf_file = {}
    conf_file["Tests"] = []
    conf_file["Cases"] = []
    conf_file["Steps"] = {}  # Will be converted to list later.
    conf_file["Reporters"] = []
    conf_file["Browsers"] = []
    wb = openpyxl.load_workbook(path, data_only=True)  # Open conf file.
    wb.name = Path(path).resolve().as_uri()

    test_run = []  # Holds all test_run worksheet names.
    test_case = []  # Holds all test_case worksheet names.
    test_reporter = []  # Holds all reporter worksheet names.
    test_browsers = []  # Holds all the browser worksheet names.

    # Get all worksheet names per type.
    for sheet in wb.sheetnames:

        test_type = wb[sheet]['A1'].value
        if test_type == 'tr_uuid:':
            test_run.append(sheet)
        elif test_type == 'tc_uuid:':
            test_case.append(sheet)
        elif test_type == 'r_uuid:':
            test_reporter.append(sheet)
        elif test_type == 'br_uuid:':
            test_browsers.append(sheet)
        else:
            print('Not a tass Excel template.')

    # Call the different convert methods for each type if they exist.

    if test_run:
        tests = convert_test_run(test_run, wb)
        breakpoint()
    if test_case:
        cases, steps = convert_test_case(test_case, conf_file, wb)
    if test_reporter:
        conf_file = convert_reporters(test_reporter, conf_file, wb)
    if test_browsers:
        browsers = convert_browsers(test_browsers, wb)

    # Convert above lists to create multiple dicts
    # each one being an execution file
    conf_file['schema-version'] = "1.0.0"
    return conf_file


def convert_browsers(browsers, wb):
    conf = []
    for browser in browsers:
        b = {} # Browser definition dict
        s = wb[browser] # Browser definition worksheet

        browser_name = s['D1'].value
        uuid = s['B1'].value

        b_args = set() # Browser arguments (flags)
        b_pref = {} # Browser preferences (key/value)
        d_config = {} # Driver configurations

        # Get Driver and browser arguments/config from sheet
        for row in s.iter_rows(min_row=3, min_col=2, max_col=6):
            if row[0].value:
                k, v = row[0].value.split(',', maxsplit=1)
                d_config[k] = v
            if row[2].value:
                b_args.add(row[2].value)
            if row[4].value:
                k, v = row[4].value.split(',', maxsplit=1)
                b_pref[k] = v

        config = {
            'driver': d_config,
            'browser': {
                'arguments': list(b_args),
                'preferences': b_pref
                }
            }

        b['browser_name'] = browser_name
        b['uuid'] = uuid
        b['configs'] = config

        conf.append(b)
    breakpoint()
    return conf


def convert_reporters(reporters, conf, wb):
    # TODO: reporters should be outputted to their own file.
    for reporter in reporters:
        r = {}
        s = wb[reporter]  # Excel sheet for reporter
        reporter_type = s['B3'].value
        if reporter_type.lower() == "testrail":
            r["uuid"] = s['B1'].value
            context = {
                "type": "testrail",
                "mode": s['B5'].value,
                "package": s['B6'].value,
                "class_name": s['B7'].value
                }
            r.update(context)
            connection = {}
            for row in s.iter_rows(min_row=5, min_col=4):
                # Connection values are comma separated key/value pairs.
                if not row:
                    continue
                k, v = row[0].value.split(",", maxsplit=1)
                connection[k] = v

            config = {}
            for row in s.iter_rows(min_row=5, min_col=6):
                # config values are comma separated key/value pairs.
                if not row:
                    continue
                k, v = row[0].value.split(",", maxsplit=1)
                config[k] = v

            r["connection"] = connection
            r["config"] = config
            conf["Reporters"].append(r)
    return conf


def convert_test_case(test_case, conf, wb):
    # TODO: Create Cases list and Steps List separately.
    for case in test_case:
        tc = {}
        tc["uuid"] = wb[case]['B1'].value
        tc["title"] = wb[case]['D1'].value
        tc["steps"] = []
        for row in wb[case].iter_rows(min_row=3, max_col=6):
            if (row[0] is None or row[0] == ''):
                continue
            print(row)
            steps = {}
            parameters = {}
            row_num = row[0].row
            steps["uuid"] = row[0].value
            steps["title"] = row[1].value
            print(row[2], row[2].value)
            steps["action"] = row[2].value.split(',', maxsplit=1)
            # steps["parameter"] = row[3].value

            # The locator value is split on the comma
            for col in wb[case].iter_cols(min_row=row_num,
                                          max_row=row_num,
                                          min_col=4):

                header = wb[case].cell(2, col[0].column).value

                if (header == '//end//'):
                    break

                if (col[0].value is not None):

                    if (header == 'locator'):
                        locator = str(col[0].value).split(',', maxsplit=1)

                        if (len(locator) == 2):
                            parameters['locator'] = {
                                'by': locator[0],
                                'value': locator[1]
                            }
                        else:
                            parameters['locator'] = locator[0]

                    elif (header == 'page'):
                        parameters['page'] = (col[0]
                                              .value
                                              .split(',', maxsplit=1))

                    elif ('action' in header):
                        parameters['action'] = (col[0]
                                                .value
                                                .split(',', maxsplit=1))

                    elif (header == 'locator_args'):
                        parameters['locator_args'] = str(col[0].value)\
                                                     .split(',')

                    elif (header == 'stored_filter'):
                        parameters['stored_filter'] = (col[0]
                                                       .value
                                                       .split(',', maxsplit=1))

                    else:
                        parameters[header] = col[0].value

            # Adds the uuid to the corresponding test case list of steps.
            tc["steps"].append(steps["uuid"])
            # This block checks to see if a step with the same uuid but
            # different values already exists. If it does, the uuid is not
            # added to the list of steps and a message is printed.
            # The title is not checked as it has no bearing on function.
            # TODO: add better way of reporting (logger? exception?)
            if row[0].value in conf["Steps"]:
                if ((conf["Steps"][row[0].value]["uuid"] == steps["uuid"])
                   and (conf["Steps"][row[0].value]["action"] ==
                        steps["action"])
                   and (conf['Steps'][row[0].value]['parameters'] ==
                        parameters)):
                    pass   # Do nothing
                else:
                    print("Different steps with uuid: " + row[0].value)
            else:
                steps['parameters'] = parameters
                conf["Steps"][row[0].value] = steps
        # Add the list of testcases to the main config file.
        conf["Cases"].append(tc)
    # Convert dictionary to list as mentioned in convert(path) above.
    conf["Steps"] = (list(conf["Steps"].values()))
    return conf


def convert_test_run(test_run, wb):
    # TODO: "Explode" tests to include configurations and cases for each test in a separate json.
    # The former Test Run object becomes a Job file.
    runs = []
    for run in test_run:
        tr = {}
        # ///// This information is converted to a Job.
        uuid = wb[run]['B1'].value
        build = wb[run]['B2'].value or "000"
        title = wb[run]['D1'].value or "TASS JOB"
        parent = wb.name

        job = {
            "uuid": uuid,
            "build": build,
            "title": title,
            "parent": parent
        }

        tr["Job"] = job
        # \\\\\

        cases = set()
        browsers = set()
        for row in wb[run].iter_rows(min_row=3, min_col=2, max_col=8):
            # Add cases
            if row[2].value:
                cases.add(row[2].value)
            # Add browsers
            if row[6].value:
                browsers.add(row[6].value)

        tests = []
        for browser in browsers:
            for c in cases:
                uuid = "--".join([job['uuid'], c, browser])
                test = {
                    "uuid": uuid,
                    "case": c,
                    "configurations": [
                        {
                            "type": "browser",
                            "uuid": browser
                        }
                    ]
                }

                tests.append(test)
        tr['Tests'] = tests

        # other attributes:
        for col in wb[run].iter_cols(min_row=3, max_row=3, min_col=10):
            if not col or col[0].value is None:
                continue
            header = col[0].value
            col_num = col[0].column
            attr = {}
            for row in wb[run].iter_rows(min_row=4,
                                         min_col=col_num,
                                         max_col=col_num):
                if not row or row[0].value is None:
                    continue
                k, v = row[0].value.split(",", maxsplit=1)
                attr[k] = v
            tr[header] = attr
        runs.append(tr)

    return runs
