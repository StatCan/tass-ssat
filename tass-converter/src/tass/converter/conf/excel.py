import openpyxl


k_V_PARAMS = ["page", "stored_filter", "locator", "offset", "delta", "target"]


def convert(path):
    """
    Helper tool to convert xlsx config file to a in-memory json equivalent.
    """
    wb = openpyxl.load_workbook(path, data_only=True)  # Open excel file.
    wb.name = path.resolve().name

    s_runs = []  # Holds all test_run worksheet names.
    s_cases = []  # Holds all test_case worksheet names.
    s_reporters = []  # Holds all reporter worksheet names.
    s_browsers = []  # Holds all the browser worksheet names.

    # Get all worksheet names per type.
    for sheet in wb.sheetnames:

        test_type = wb[sheet]['A1'].value
        if test_type == 'tr_uuid:':
            s_runs.append(sheet)
        elif test_type == 'tc_uuid:':
            s_cases.append(sheet)
        elif test_type == 'r_uuid:':
            s_reporters.append(sheet)
        elif test_type == 'br_uuid:':
            s_browsers.append(sheet)
        else:
            print('Not a tass Excel template.')

    # Extract configurations from sheets using sheet names
    if s_cases:
        # Extract cases and steps from sheets
        cases, steps = convert_cases(s_cases, wb)
    if s_browsers:
        # Extract browser configs from sheets
        browsers = convert_browsers(s_browsers, wb)
    if s_runs:
        # Extract all jobs to executed as a list.
        runs = convert_runs(s_runs, cases, steps, browsers, wb)
    if s_reporters:
        # TODO: define and print reporters.
        reporters = convert_reporters(s_reporters, wb)

    return runs


def convert_cases(cases, wb):
    _cases = {}  # Dict of all test cases. Ensures unique case ids.
    _steps = {}  # Dict of all test steps. Ensures unique step ids.

    for c in cases:
        tc = {}  # Test case dict, to be printed in JSON
        tc["uuid"] = wb[c]['B1'].value

        if tc['uuid'] in _cases:
            print('UUID already exists. Skipping this case.')
            continue
        tc["title"] = wb[c]['D1'].value
        tc["steps"] = []

        for row in wb[c].iter_rows(min_row=3, max_col=6):
            if (row[0] is None or row[0] == ''):
                continue
            steps = {}
            parameters = {}
            row_num = row[0].row
            valid = all(
                [row[0].value,
                isinstance(row[0].value, str),
                row[1].value,
                isinstance(row[1].value, str),
                row[2].value,
                isinstance(row[2].value, str),
                "," in str(row[2].value)]
            )

            if not valid:
                print(f"Case: {tc["uuid"]} Row: {row_num} Step is missing mandatory fields UUID, TITLE, or ACTION.")
                continue
            steps["uuid"] = row[0].value
            steps["title"] = row[1].value
            steps["action"] = row[2].value.split(',', maxsplit=1)

            # The locator value is split on the comma
            for col in wb[c].iter_cols(min_row=row_num,
                                       max_row=row_num,
                                       min_col=4):

                header = wb[c].cell(2, col[0].column).value

                if (header == '//end//'):  # e.o.f indicator for TASS
                    break

                header = header.lower()

                if (col[0].value is not None):

                    # locator parameter.
                    # key,value pair or single string.
                    # ',' is reserved character.
                    if header in k_V_PARAMS:
                        param = str(col[0].value).split(',', maxsplit=1)

                        if header == 'locator' or header == 'target':
                            if (len(param) == 2):
                                parameters[header] = {
                                    'by': param[0],
                                    'value': param[1]
                                }
                            else:
                                parameters[header] = param[0]
                        elif header == 'offset':
                            parameters['xoffset'] = int(param[0])
                            parameters['yoffset'] = int(param[1])
                        elif header == 'delta':
                            parameters['deltax'] = int(param[0])
                            parameters['deltay'] = int(param[1])
                        else:
                            parameters[header] = param

                    # action parameter. read as "module,action"
                    elif ('action' in header):
                        parameters['action'] = (col[0]
                                                .value
                                                .split(',', maxsplit=1))

                    elif (header == 'locator_args'):
                        # locator_args parameter.
                        # "," separated list of arguments for locator.
                        # arguments are filled in order
                        # of definition at run time.
                        parameters[header] = (str(col[0].value)
                                              .split(','))

                    else:
                        parameters[header] = col[0].value

            # Adds the uuid to the corresponding test case list of steps.
            tc["steps"].append(steps["uuid"])
            # This block checks to see if a step with the same uuid but
            # different values already exists. If it does, the uuid is not
            # added to the list of steps and a message is printed.
            # The title is not checked as it has no bearing on function.
            # TODO: add better way of reporting (logger? exception?)
            if row[0].value in _steps:
                if ((_steps[row[0].value]["uuid"] == steps["uuid"])
                    and (_steps[row[0].value]["action"] ==
                         steps["action"])
                    and (_steps[row[0].value]['parameters'] ==
                         parameters)):
                    pass   # Do nothing
                else:
                    print("Different steps with uuid: " + row[0].value)
            else:
                steps['parameters'] = parameters
                _steps[row[0].value] = steps

        _cases[tc['uuid']] = tc
    return _cases, _steps


def convert_browsers(browsers, wb):
    conf = {}
    for browser in browsers:
        b = {}  # Browser definition dict
        s = wb[browser]  # Browser definition worksheet

        browser_name = s['D1'].value
        uuid = s['B1'].value

        b_args = set()  # Browser arguments (flags)
        b_pref = {}  # Browser preferences (key/value)
        d_config = {}  # Driver configurations

        # Get Driver and browser arguments/config from sheet
        for row in s.iter_rows(min_row=3, min_col=2, max_col=6):
            if row[0].value:
                k, v = row[0].value.split(',', maxsplit=1)
                d_config[k] = v
            if row[2].value:
                b_args.add(row[2].value)
            if row[4].value:
                k, v = row[4].value.split(',', maxsplit=1)
                b_pref[k] = v

        config = {
            'driver': d_config,
            'browser': {
                'arguments': sorted(list(b_args)),
                'preferences': b_pref
                }
            }

        b['browser_name'] = browser_name
        b['uuid'] = uuid
        b['configs'] = config

        conf[uuid] = b
    return conf


def convert_runs(runs, cases, steps, browsers, wb):
    execs = []  # Holds all ready-to-run test runs as dict

    for run in runs:
        tr = {}  # Test run to be built
        ts = wb[run]  # Excel sheet containing test run configs.
        # ///// This information is converted to a Job.
        uuid = ts['B1'].value
        build = ts['B2'].value or "000"
        title = ts['D1'].value or "TASS JOB"
        parent = wb.name

        job = {
            "uuid": uuid,
            "build": build,
            "title": title,
            "parent": parent
            }

        tr["Job"] = job
        tr["schema-version"] = "1.1.0"
        # \\\\\

        # ///// Create custom Logger as applicable
        if ts["D2"].value:
            logger_path = ts["D2"].value
            tr["Logger"] = {"path": logger_path}
        # \\\\\

        caseset = set()
        browserset = set()
        stepset = set()
        for row in wb[run].iter_rows(min_row=3, min_col=2, max_col=8):
            # Add cases ids
            if row[2].value:
                caseset.add(row[2].value)
            # Add browsers ids
            if row[6].value:
                browserset.add(row[6].value)

        for case in caseset:
            # Add step ids
            c = cases[case]
            for step in c['steps']:
                stepset.add(step)

        # Extract cases for this job
        tr["Cases"] = sorted([cases[c] for c in caseset],
                             key=lambda x: x['uuid'])
        # Extract steps for this job
        tr["Steps"] = sorted([steps[s] for s in stepset],
                             key=lambda x: x['uuid'])
        # Extract browsers for this job
        tr["Browsers"] = sorted([browsers[b] for b in browserset],
                                key=lambda x: x['uuid'])

        tests = []
        # "Explode" test cases + configurations
        for browser in browserset:
            for c in caseset:
                uuid = "--".join([job['uuid'], c, browser])  # Hybrid uuid
                test = {
                    "uuid": uuid,
                    "case": c,
                    "configurations": [
                        {
                            "type": "browser",
                            "uuid": browser
                        }
                    ]
                }

                tests.append(test)
        tr['Tests'] = sorted(tests, key=lambda x: x['uuid'])

        # other attributes:
        for col in wb[run].iter_cols(min_row=3, max_row=3, min_col=10):
            if not col or col[0].value is None:
                continue
            header = col[0].value
            col_num = col[0].column
            attr = {}
            for row in wb[run].iter_rows(min_row=4,
                                         min_col=col_num,
                                         max_col=col_num):
                if not row or row[0].value is None:
                    continue
                k, v = row[0].value.split(",", maxsplit=1)
                attr[k] = v
            tr[header] = attr

        execs.append(tr)

    return execs


def convert_reporters(reporters, wb):
    # TODO: define reporters
    pass
