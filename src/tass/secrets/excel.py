import openpyxl
import tass.secrets.secrets as secrets
from pathlib import Path
from collections import namedtuple


class Excel(secrets.DataSource):

    def __init__(self, config_path, config):
        super().__init__(config_path, config)
    

    def _load_datasource(self, config):
        self._source_path = Path(config['source']['path']).resolve()
        _source = config['source']
        _all_collections = config['collections']

        _collections = [_c for _c in _all_collections if _c['name'] in _source['collections']] 


        self._collections = self._load_collections(
                                openpyxl.load_workbook(self._source_path),
                                _collections, config['entry-sets'])
        
        
    def _load_collections(self, source, collections, entry_sets):
        _loaded = {}
        for coll in collections:
            _name = coll['name']
            _entry_set = [se for se in entry_sets if se['name'] == coll['entry-set']][0]
            _wsheet = source[_name]
            _loaded[_name] = Excel.Sheet(self, _wsheet, _entry_set)
        return _loaded

    def save_changes(self):
        if not self._changed:
                return
        wb = openpyxl.load_workbook(self._source_path)
        for collection in self._collections.values():
            if not collection._changed:
                continue
            ws = wb[collection._name]
            for entry in collection._entries.values():
                if entry.changed:
                    row = entry.get('row')
                    for column in collection._columns:
                        col = column.column + 1
                        cell = ws.cell(row=row, column=col)
                        cell.value = entry.get(column.name)
        
        wb.save(self._source_path)

            
        

    class Sheet(secrets.Collection):
        def __init__(self, parent, collection, entry_set):
            super().__init__(parent, collection=collection, entry_set=entry_set)
            

        def _load_entries(self, collection, entry_set):
            self._columns = []
            self._name = collection.title
            _key = entry_set['key']
            _cols = entry_set['columns']
            _has_headers = entry_set['has-headers']
            ColumnDefinition = namedtuple("ColumnDefinition", "name, column")
            for col in _cols:
                self._columns.append(ColumnDefinition(col['name'], col['column']))
            
            start_row = 2 if _has_headers else 1
            _entries = {}
            for index, row in enumerate(collection.iter_rows(min_row=start_row), start_row):
                entry = {"row": index}
                for col in self._columns:
                    entry[col.name] = row[col.column].value
                _entries[entry[_key]] = Excel.Row(self, entry)
            
            return _entries

        @property
        def columns(self):
            return self._columns
            

    class Row(secrets.Entry):
        def __init__(self, sheet, data):
            super().__init__(sheet, data)