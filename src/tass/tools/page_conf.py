import openpyxl
from pathlib import Path
from enum import StrEnum, auto


class Element(StrEnum):
    SECTION = auto()
    ANSWER = auto()
    ROSTER = auto()
    FIELD = auto()
    BREAK = "pagebreak"
    GROUP = "groupheader"
    ENDGROUP = 'endgroupheader'


class ElementType(StrEnum):
    RADIO = auto()
    RADIOTEXT = auto()
    CHECK = auto()
    CHECKSPECIFY = auto()
    DROPDOWN = auto()
    INTEGER = auto()
    TEXT = auto()
    PHONE = auto()
    EMAIL = auto()
    COMMENTBOX = auto()
    POSTALNOVAL = "postalcodenoval"
    INFO = auto()


def convert_to_excel(specs_path):
    """
    Helper tool to convert EQ specs to Page model
    """

    wb = openpyxl.load_workbook(specs_path)  # Open the specs

    spec_name = "Capture Specifications"
    if spec_name not in wb.sheetnames:
        return  "Specs not found"  # Throw some exception?

    wb_out = openpyxl.Workbook()
    specs_out = wb_out.active
    specs_out.title = "Pages"

    specs_out.append(["Element Type", "ID", "Rostered", "Title", "URL"])
    specs_out.append(["pagebreak"])

    specs = wb[spec_name]

    def parse_sheet(sheet, specs_out):

        ele = []
        hidden = False

        field_type = None
        radio_group = None

        for row in sheet.iter_rows(min_row=2):
            row_type = row[0].value.lower()
            if row_type == Element.BREAK:
                hidden = False
                field_type = None
                radio_group = None
                if len(ele) > 0:
                    for e in ele:
                        print(e)
                        specs_out.append(e)

                    specs_out.append(["pagebreak"])
                    ele = []

            elif hidden:
                print("Row is hidden:", row[0].row)
                continue
            else:

                if row_type == Element.SECTION:
                    if "hidden" in row[7].value.lower():
                        hidden = True

                elif row_type == Element.ANSWER:
                    field_type = None
                    element_type = row[1].value.lower()
                    print("This is an anser. It is:", element_type)
                    match element_type:
                        case ElementType.RADIOTEXT:
                            print(ElementType.RADIOTEXT)
                            field_type = ElementType.RADIOTEXT
                        case ElementType.RADIO:
                            print(ElementType.RADIO)
                            field_type = ElementType.RADIO
                            radio_group = row[2].value
                        case (ElementType.TEXT |
                                ElementType.INTEGER |
                                ElementType.PHONE |
                                ElementType.EMAIL |
                                ElementType.POSTALNOVAL |
                                ElementType.COMMENTBOX):
                            ele.append([ElementType.TEXT, row[2].value])
                        case ElementType.DROPDOWN:
                            ele.append([ElementType.DROPDOWN, row[2].value])
                        case ElementType.CHECK | ElementType.CHECKSPECIFY:
                            ele.append([ElementType.CHECK, row[2].value])
                        case ElementType.INFO:
                            ele.append([ElementType.INFO, row[2].value])

                elif row_type == Element.FIELD:
                    print("This row is a field: ", row[0].row)
                    print("The field is of the type: ", field_type)
                    if field_type == ElementType.RADIOTEXT:
                        ele.append([ElementType.RADIOTEXT, row[5].value])

                    elif field_type == ElementType.RADIO:
                        _id = f"{radio_group}-{row[5].value}"
                        ele.append([ElementType.RADIO, _id])

        if len(ele) > 0:
            for e in ele:
                specs_out.append(e)

    parse_sheet(specs, specs_out)

    submit_page_name = "Submit Page"
    if (submit_page_name in wb.sheetnames):
        specs = wb[submit_page_name]
        parse_sheet(specs, specs_out)

    wb_out.save("pages.xlsx")
    
    return str(Path("specs.xlsx").resolve())


def convert_to_json(pages_path):
    """ 
    Convert the results of convert to framework json format
    """
    
    wb = openpyxl.load_workbook(pages_path)
    
    if ("Pages" in wb.sheetnames):
        return  # TODO: Throw exception??
    ws = wb["Pages"]
    
    pages = {}
    page = {}
    elements = {}
    page_number = 0
    def parse_radio(id, val, rostered):
        if rostered:
            return f'//input[contains(@name, ".Instance") and @value={{}}]/following-sibling::*/descendant::input[contains(@name, ".{id}") and @value={val}]'
        else:
            return f'//input[contains(@name, ".{id}") and @value={val}]'
    
    def parse_radiotext(id, rostered):
        if rostered:
            return f'//input[contains(@name, ".Instance") and @value={{}}]/following-sibling::*/descendant::input[@value="{id}"]'
        else:
            return f'//input[@value="{id}"]'
    
    def parse_check(id, rostered):
        if rostered:
            return f'//input[contains(@name, ".Instance") and @value={{}}]/following-sibling::*/descendant::input[contains(@name, ".{id}") and @type="checkbox"]'
        else:
            return f'//input[@name="{id}" and @type="checkbox"]'
    
    def parse_dropdown(id, rostered):
        if rostered:
            return f'//input[contains(@name, ".Instance") and @value={{}}]/following-sibling::*/descendant::select[contains(@name, ".{id}")]'
        else:
            return f'//select[@name="{id}"]'
    
    def parse_text(id, rostered):
        if rostered:
            return f'//input[contains(@name, ".Instance") and @value={{}}]/following-sibling::*/descendant::input[contains(@name, ".{id}")]'
        else:
            return f'//input[@name="{id}"]'
    
    for row in ws.iter_rows(min_row=2, max_col=5):
        row_type = row[0].value
        element_id = row[1].value
        rostered = row[2].value
        if row_type == 'pagebreak':
            # TODO: start new page, add existing page to pages
            # TODO: add page identity to page using id
            if not page and not elements:

                if row[1].value:
                    page_number = row[1].value
                else:
                    page_number += 1
                
            else:
                page["elements"] = elements
                pages[f"p{page_number}"] = page
                
                page = {}
                elements= {}
                
            page['title'] = row[3].value
            page['url'] = row[4].value
            page['page_id'] = {"method": "element", "identifier": {"by": "xpath", "value": f"//*[@id='__pageId' and @value='p{page_number}']"}}
            
        elif row_type == ElementType.INFO:
            # TODO: XPATH for info fields
        elif row_type == ElementType.CHECK:
            elements[element_id] = {"by": "xpath", "value": parse_check(element_id, rostered)}
        elif row_type == ElementType.RADIO:
            _id, _val = element_id.split("-")
            elements[element_id] = {"by": "xpath", "value": parse_radio(_id, _val, rostered)}
        elif row_type == ElementType.RADIOTEXT:
            elements[element_id] = {"by": "xpath", "value": parse_radiotext(element_id, rostered)}
        elif row_type == ElementType.TEXT:
            elements[element_id] = {"by": "xpath", "value": parse_text(element_id, rostered)}
        elif row_type == ElementType.DROPDOWN:
            elements[element_id] = {"by": "xpath", "value": parse_dropdown(element_id, rostered)}
        else:
            continue
            # TODO: Fallback condition. Error?