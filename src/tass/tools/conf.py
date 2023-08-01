import openpyxl


def convert(path):
    """
    Helper tool to convert xlsx config file to a in-memory json equivalent.
    """
    conf_file = {}
    conf_file["Test_runs"] = []
    conf_file["Test_suites"] = []
    conf_file["Test_cases"] = []
    conf_file["Steps"] = {}  # Will be converted to list later.
    wb = openpyxl.load_workbook(path)  # Open conf file.

    test_run = []  # Holds all test_run worksheet names.
    test_suite = []  # Holds all test_suite worksheet names.
    test_case = []  # Holds all test_case worksheet names.

    # Get all worksheet names per type.
    for sheet in wb.sheetnames:

        test_type = wb[sheet]['A1'].value
        if test_type == 'tr_uuid:':
            test_run.append(sheet)
        elif test_type == 'ts_uuid:':
            test_suite.append(sheet)
        elif test_type == 'tc_uuid:':
            test_case.append(sheet)
        else:
            print('Not a tass Excel template.')

    # Call the different convert methods for each type if they exist.
    if test_run:
        conf_file = convert_test_run(test_run, conf_file, wb)
    if test_suite:
        conf_file = convert_test_suite(test_suite, conf_file, wb)
    if test_case:
        conf_file = convert_test_case(test_case, conf_file, wb)

    return conf_file


def convert_test_case(test_case, conf, wb):
    conf["Test_cases"] = []
    for case in test_case:
        tc = {}
        tc["uuid"] = wb[case]['B1'].value
        tc["title"] = wb[case]['D1'].value
        tc["steps"] = []
        for row in wb[case].iter_rows(min_row=3, max_col=6):
            print(row)
            steps = {}
            parameters = {}
            row_num = row[0].row
            steps["uuid"] = row[0].value
            steps["title"] = row[1].value
            print(row[2], row[2].value)
            steps["action"] = row[2].value.split(',', 1)
            # steps["parameter"] = row[3].value

            # The locator value is split on the comma
            if (row[3].value is not None):
                locator = row[3].value.split(',')

                if (len(locator) == 2):
                    parameters['locator'] = {
                        'by': locator[0],
                        'value': locator[1]
                    }
                else:
                    parameters['locator'] = locator[0]

            if (row[4].value is not None):
                parameters['page'] = row[4].value.split(',', 1)

            if (row[5].value is not None):
                parameters['action'] = row[5].value.split(',', 1)

            for col in wb[case].iter_cols(min_row=row_num,
                                          max_row=row_num,
                                          min_col=7):
                if (col[0].value is not None):
                    header = wb[case].cell(2, col[0].column).value
                    parameters[header] = col[0].value

            # Adds the uuid to the corresponding test case list of steps.
            tc["steps"].append(steps["uuid"])
            # This block checks to see if a step with the same uuid but
            # different values already exists. If it does, the uuid is not
            # added to the list of steps and a message is printed.
            # TODO: add better way of reporting (logger? exception?)
            if row[0].value in conf["Steps"]:
                if ((conf["Steps"][row[0].value]["uuid"] == steps["uuid"])
                   and (conf["Steps"][row[0].value]["title"] == steps["title"])
                   and (conf["Steps"][row[0].value]["action"] ==
                        steps["action"])
                   and (conf['Steps'][row[0].value]['parameters'] ==
                        parameters)):
                    pass   # Do nothing
                else:
                    print("Different steps with uuid: " + row[0].value)
            else:
                steps['parameters'] = parameters
                conf["Steps"][row[0].value] = steps
        # Add the list of testcases to the main config file.
        conf["Test_cases"].append(tc)
    # Convert dictionary to list as mentioned in convert(path) above.
    conf["Steps"] = (list(conf["Steps"].values()))
    return conf


def convert_test_suite(test_suite, conf, wb):
    conf["Test_suites"] = []
    for suite in test_suite:
        ts = {}
        ts["uuid"] = wb[suite]['B1'].value
        ts["test_cases"] = []
        ts["keywords"] = []
        for row in wb[suite].iter_rows(min_row=2, min_col=2, max_col=2):
            if row[0].value is not None:
                ts["test_cases"].append(row[0].value)
        for row in wb[suite].iter_rows(min_row=1, min_col=4, max_col=4):
            if row[0].value is not None:
                ts["keywords"].append(row[0].value)
        conf["Test_suites"].append(ts)

    return conf


def convert_test_run(test_run, conf, wb):
    conf["Test_runs"] = []
    for run in test_run:
        tr = {}
        tr["uuid"] = wb[run]['B1'].value
        tr["build"] = wb[run]['B2'].value
        # tr["start_time"] = wb[run]['D1'].value
        # tr["end_time"] = wb[run]['D2'].value
        tr["test_cases"] = []
        tr["test_suites"] = []
        for row in wb[run].iter_rows(min_row=3, min_col=2, max_col=2):
            if row[0].value is not None:
                tr["test_suites"].append(row[0].value)
        for row in wb[run].iter_rows(min_row=3, min_col=4, max_col=4):
            if row[0].value is not None:
                tr["test_cases"].append(row[0].value)
        conf["Test_runs"].append(tr)

    return conf
